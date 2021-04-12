import openpyxl

# print("hello world")

# a = input("请输入第一个值：")
# b = input("请输入第二个值：")
# a = int(a)
# b = int(b)
# c = a / b
# print("%d与%d相除等于%d" % (a, b, c))
# print(type(a))

# l = [1, 2, 3, 4]

# print(type(l), l[3], l[:4:3])

# for i in l[1:4:2]:
#     print(i)
# num = 0
# while num < len(l):
#     print(num,l[num])
#     num=num+1
# l = [1, 2, 3, 4, 5, 6, 7, 8, 9]
#
# for i in l:
#     for j in l[i - 1:]:
#         m = i * j
#         print('%d*%d=%d' % (i, j, m))


# def nn():
#     num = 0
#     while num < 9:
#         num = num + 1
#         k = num
#         while k < 10:
#             r = num * k
#             print('%d*%d=%d' % (num, k, r), end=" ")
#             k = k + 1
#         print("")
#
# import main
#
# main.nn()


# workbook = openpyxl.load_workbook("万联诊股升级文档.xlsx")
# worksheet = workbook.active
# print(worksheet["D3"].value)
# worksheet['E1'].value = "总分"
# print(worksheet.cell(1,2).value,worksheet.iter_rows())
# for r in worksheet.iter_rows(min_row=2):
#     end = r[2].value + r[3].value
#     end = end / 2
#     r[5].value = end
#     r[4].value = "=sum()
# worksheet['F1'].value = "平均分"

import openpyxl

workbook = openpyxl.load_workbook("万联诊股升级文档.xlsx")
workbook.remove(workbook["Sheet2"])
workbook.create_sheet(title='Sheet2')
worksheet = workbook.active
for i in range(1, 10):
    for j in range(i, 10):
        end = i * j
        worksheet.cell(row=i, column=j).value = "%d*%d=%d" % (i, j, end)

# worksheet['A1'].value = n
workbook.save("万联诊股升级文档.xlsx")
